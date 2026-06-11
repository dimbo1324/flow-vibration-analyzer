"""Unit tests for iva.infrastructure.readers.excel_reader."""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest

pytest.importorskip("openpyxl", reason="openpyxl is required for Excel tests")

import openpyxl  # noqa: E402

from iva.core.models.exceptions import FileReadError, IVAFileNotFoundError  # noqa: E402
from iva.infrastructure.readers.excel_reader import read_excel  # noqa: E402


def _write_xlsx(headers: list[str], rows: list[list]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    fd, name = tempfile.mkstemp(suffix=".xlsx")
    import os

    os.close(fd)
    p = Path(name)
    wb.save(p)
    wb.close()
    return p


def test_read_excel_basic():
    """A well-formed xlsx file with a header row is parsed correctly."""
    p = _write_xlsx(["time_s", "signal"], [[0.0, 1.0], [0.001, 1.1], [0.002, 0.9]])
    try:
        result = read_excel(str(p))
        assert result.file_format == "xlsx"
        assert result.row_count == 3
        assert "time_s" in result.column_names
        assert isinstance(result.file_path, Path)
    finally:
        p.unlink(missing_ok=True)


def test_read_excel_explicit_sheet_name():
    """Passing an explicit sheet name reads the correct sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["time_s", "signal"])
    ws.append([0.0, 1.0])
    fd, name = tempfile.mkstemp(suffix=".xlsx")
    import os

    os.close(fd)
    p = Path(name)
    wb.save(p)
    wb.close()
    try:
        result = read_excel(str(p), sheet_name="Data")
        assert result.row_count == 1
    finally:
        p.unlink(missing_ok=True)


def test_read_excel_missing_sheet_raises():
    """FileReadError is raised when the requested sheet does not exist."""
    p = _write_xlsx(["time_s", "signal"], [[0.0, 1.0]])
    try:
        with pytest.raises(FileReadError):
            read_excel(str(p), sheet_name="NoSuchSheet")
    finally:
        p.unlink(missing_ok=True)


def test_read_excel_file_not_found():
    """IVAFileNotFoundError is raised for a missing xlsx file."""
    with pytest.raises(IVAFileNotFoundError):
        read_excel("/nonexistent/file.xlsx")


def test_read_excel_preserves_literal_none_string():
    """A cell whose text is literally 'None' is kept, while blank cells become NA.

    Regression test: an earlier implementation stringified every column and then
    replaced the string 'None' with NA, which destroyed legitimate textual
    'None' values.  The reader must distinguish a real 'None' string from an
    empty cell.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["label", "signal"])
    ws.append(["None", 1.0])  # literal text "None" — must be preserved
    ws.append([None, 2.0])  # genuinely empty cell — must become NA
    fd, name = tempfile.mkstemp(suffix=".xlsx")
    import os

    os.close(fd)
    p = Path(name)
    wb.save(p)
    wb.close()
    try:
        result = read_excel(str(p))
        labels = result.data["label"]
        assert labels.iloc[0] == "None", "literal 'None' string must be preserved"
        assert labels.isna().iloc[1], "empty cell must be NA, not the string 'None'"
    finally:
        p.unlink(missing_ok=True)


def test_read_excel_read_only_mode():
    """openpyxl is called with read_only=True (security requirement).

    The read_only constraint is enforced in the reader source.  This test
    verifies the constraint by patching load_workbook and inspecting kwargs.
    """
    from unittest.mock import patch

    captured: list[dict] = []
    real_load = openpyxl.load_workbook

    def _spy(*args, **kwargs):
        captured.append(kwargs)
        return real_load(*args, **kwargs)

    p = _write_xlsx(["time_s", "signal"], [[0.0, 1.0]])
    try:
        with patch("openpyxl.load_workbook", side_effect=_spy):
            result = read_excel(str(p))
        assert result.row_count == 1
        assert captured, "load_workbook was never called"
        assert captured[0].get("read_only") is True
    finally:
        p.unlink(missing_ok=True)
