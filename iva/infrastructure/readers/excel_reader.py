"""Безопасное чтение Excel (.xlsx) для IVA.

``openpyxl`` открывает книгу в ``read_only=True`` и ``data_only=True`` согласно
docs/18_security_and_data_privacy.md: макросы не исполняются, формулы читаются
как сохранённые значения.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

from iva.core.models.exceptions import FileCorruptedError, FileReadError, IVAFileNotFoundError
from iva.core.models.signal_data import RawFileData
from iva.infrastructure.logging.app_logger import get_logger

logger = get_logger(__name__)


def _is_header_row(row: tuple[Any, ...]) -> bool:
    """Проверить, выглядит ли строка как текстовый заголовок."""
    has_string = False
    for cell in row:
        if cell is None:
            continue
        try:
            float(str(cell))
            return False  # Числовое значение указывает на строку данных.
        except (ValueError, TypeError):
            has_string = True
    return has_string


def read_excel(file_path: str, sheet_name: str | None = None) -> RawFileData:
    """Прочитать лист Excel и вернуть ``RawFileData``.

    Opens the workbook in ``read_only=True, data_only=True`` mode (macros are
    never executed; formula results are read as cached values).

    Args:
        file_path: Absolute or relative path to the ``.xlsx`` file.
        sheet_name: Name of the worksheet to read.  If ``None``, the first
            worksheet is used.

    Returns:
        A populated :class:`RawFileData` instance.

    Raises:
        IVAFileNotFoundError: The file does not exist.
        FileReadError: The file exists but cannot be parsed, or the requested
            sheet does not exist.
    """
    # Ленивый импорт сохраняет возможность импортировать модуль без openpyxl.
    import openpyxl

    path = Path(file_path)
    logger.debug("Excel read started: %s", path.name)

    if not path.exists():
        raise IVAFileNotFoundError(
            user_message="Входной файл не найден.",
            technical_details=f"Full path checked: {path}",
            recovery_hint="Проверьте путь к файлу и убедитесь, что файл не был перемещен.",
        )

    file_size = path.stat().st_size

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except FileNotFoundError as exc:
        raise IVAFileNotFoundError(
            user_message="Входной файл не найден.",
            technical_details=str(exc),
        ) from exc
    except Exception as exc:
        raise FileCorruptedError(
            user_message="Не удалось открыть файл Excel. Возможно, файл поврежден "
            "или имеет неподдерживаемый формат.",
            technical_details=str(exc),
            recovery_hint="Повторно сохраните файл в формате .xlsx и попробуйте снова.",
        ) from exc

    try:
        if sheet_name is None:
            ws = wb.worksheets[0]
            sheet_name = ws.title
        else:
            if sheet_name not in wb.sheetnames:
                raise FileReadError(
                    user_message=f"Лист '{sheet_name}' не найден в книге Excel.",
                    technical_details=f"Available sheets: {wb.sheetnames}",
                    recovery_hint="Проверьте имя листа и повторите попытку.",
                )
            ws = wb[sheet_name]

        rows: list[tuple[Any, ...]] = [tuple(cell.value for cell in row) for row in ws.iter_rows()]
    finally:
        wb.close()

    if not rows:
        raise FileReadError(
            user_message="Выбранный лист Excel пуст.",
            technical_details=f"Sheet '{sheet_name}' contained no rows.",
        )

    # Заголовок определяется эвристически, но имена ролей всё равно назначает пользователь.
    first_row = rows[0]
    if _is_header_row(first_row):
        headers = [str(v) if v is not None else f"column_{i}" for i, v in enumerate(first_row)]
        data_rows = rows[1:]
    else:
        n_cols = len(first_row)
        headers = [f"column_{i}" for i in range(n_cols)]
        data_rows = rows

    try:
        df = pd.DataFrame(data_rows, columns=headers)
    except Exception as exc:
        raise FileReadError(
            user_message="Не удалось преобразовать лист Excel в таблицу.",
            technical_details=str(exc),
        ) from exc

    # Типы нормализуются к строкам как в CSV-reader. Настоящие пустые ячейки
    # восстанавливаются как pd.NA после преобразования, чтобы текст ``"None"``
    # не был ошибочно принят за пропуск.
    for col in df.columns:
        series = df[col]
        is_null = series.isna()
        df[col] = series.astype(str)
        df.loc[is_null, col] = pd.NA

    column_names = tuple(str(c) for c in df.columns)
    column_dtypes = {col: str(df[col].dtype) for col in df.columns}
    row_count = len(df)

    logger.info(
        "Excel read completed: '%s' sheet='%s' — %d rows, %d columns",
        path.name,
        sheet_name,
        row_count,
        len(column_names),
    )

    return RawFileData(
        file_path=path,
        file_format="xlsx",
        column_names=column_names,
        column_dtypes=column_dtypes,
        row_count=row_count,
        file_size_bytes=file_size,
        data=df,
        read_timestamp=datetime.now(),
    )
